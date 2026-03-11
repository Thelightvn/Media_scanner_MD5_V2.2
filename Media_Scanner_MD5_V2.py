import os
import subprocess
import sys
import shutil
import hashlib
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime
from collections import defaultdict

# ==============================================================================
# PLATFORM DETECTION
# ==============================================================================
IS_WINDOWS = sys.platform == "win32"
IS_MACOS   = sys.platform == "darwin"


# ==============================================================================
# PHẦN LOGIC XỬ LÝ
# ==============================================================================

MEDIA_EXTENSIONS = (
    '.mp4', '.mkv', '.avi', '.mov', '.wmv', '.flv', '.webm', '.mpg', '.mpeg', '.mxf',
    '.mp3', '.wav', '.flac', '.aac', '.ogg', '.m4a'
)

DPX_CRI_EXTENSIONS = ('.dpx', '.cri')

ALL_EXTENSIONS = MEDIA_EXTENSIONS + DPX_CRI_EXTENSIONS


def format_size(size_bytes):
    if size_bytes == 0: return "0B"
    size_name = ("B", "KB", "MB", "GB", "TB")
    i = 0
    while size_bytes >= 1024 and i < len(size_name) - 1:
        size_bytes /= 1024.0
        i += 1
    return f"{size_bytes:.2f} {size_name[i]}"


def format_duration(seconds):
    if seconds is None: return "N/A"
    try:
        seconds = float(seconds)
        hours = int(seconds // 3600)
        minutes = int((seconds % 3600) // 60)
        secs = int(seconds % 60)
        return f"{hours:02}:{minutes:02}:{secs:02}"
    except (ValueError, TypeError):
        return "N/A"


def calculate_md5(file_path, chunk_size=8 * 1024 * 1024):
    """
    Tính MD5 checksum của file.
    chunk_size=8MB: tối ưu cho HDD sequential read, giảm số lần syscall.
    Trả về chuỗi hex 32 ký tự, hoặc chuỗi lỗi nếu thất bại.
    """
    md5 = hashlib.md5()
    try:
        with open(file_path, 'rb', buffering=0) as f:
            # buffering=0 → bypass Python buffer, đọc thẳng từ OS cache
            mv = memoryview(bytearray(chunk_size))
            while True:
                n = f.readinto(mv)
                if not n:
                    break
                md5.update(mv[:n])
        return md5.hexdigest().upper()
    except PermissionError:
        return "ERR:NO_PERMISSION"
    except Exception as e:
        return f"ERR:{e}"


def get_ffprobe_path():
    """
    Tìm ffprobe theo thứ tự ưu tiên:
    1. Cùng thư mục với exe (bundled bởi PyInstaller)
    2. Thư mục _internal (PyInstaller --onedir)
    3. sys._MEIPASS (PyInstaller --onefile)
    4. PATH của hệ thống
    """
    ffprobe_name = "ffprobe.exe" if IS_WINDOWS else "ffprobe"

    # Thư mục chứa exe đang chạy
    if getattr(sys, 'frozen', False):
        # Đang chạy từ PyInstaller bundle
        base_dir = os.path.dirname(sys.executable)
        candidates = [
            os.path.join(base_dir, ffprobe_name),                        # cùng thư mục exe
            os.path.join(base_dir, "_internal", ffprobe_name),           # thư mục _internal
            os.path.join(getattr(sys, '_MEIPASS', base_dir), ffprobe_name),  # onefile temp
        ]
    else:
        # Đang chạy từ source .py
        base_dir = os.path.dirname(os.path.abspath(__file__))
        candidates = [
            os.path.join(base_dir, ffprobe_name),   # cùng thư mục script
        ]

    for path in candidates:
        if os.path.isfile(path):
            return path

    # Fallback: tìm trong PATH hệ thống
    return shutil.which("ffprobe") or shutil.which("ffprobe.exe")


def get_media_duration(file_path):
    """Lấy thời lượng file media. Trả về None cho DPX/CRI vì xử lý riêng."""
    ext = os.path.splitext(file_path)[1].lower()
    if ext in DPX_CRI_EXTENSIONS:
        return None

    ffprobe = get_ffprobe_path()
    if not ffprobe:
        return None

    command = [
        ffprobe, "-v", "error", "-show_entries", "format=duration",
        "-of", "default=noprint_wrappers=1:nokey=1", file_path
    ]
    try:
        kwargs = dict(
            stdout=subprocess.PIPE, stderr=subprocess.PIPE,
            text=True, check=False, encoding='utf-8', errors='ignore'
        )
        if IS_WINDOWS:
            si = subprocess.STARTUPINFO()
            si.dwFlags |= subprocess.STARTF_USESHOWWINDOW
            kwargs['startupinfo'] = si

        result = subprocess.run(command, **kwargs)
        if result.returncode != 0:
            return None
        return result.stdout.strip()

    except FileNotFoundError:
        return None
    except Exception as e:
        print(f"An unexpected error occurred in get_media_duration: {e}")
        return None


def check_dpx_cri_file(file_path):
    """
    Kiểm tra header và tính hợp lệ của file DPX hoặc CRI.

    DPX (Digital Picture Exchange - SMPTE 268M):
      - Magic BE: 0x53445058 ("SDPX")
      - Magic LE: 0x58504453 ("XPDS")
      - Header 2048 bytes, chứa width/height/bit-depth/colorspace
      - Hỗ trợ 8/10/12/16-bit, RGB/YCbCr/grayscale

    CRI Middleware (phổ biến trong game):
      - USM video  : magic "CRID" (0x43524944)
      - HCA audio  : magic "HCA\x00" (0x48434100)
      - ADX audio  : magic 0x80 0x00
      - UTF table  : magic "@UTF" (0x40555446)
    """
    ext = os.path.splitext(file_path)[1].lower()
    result = {
        "type": ext.upper().lstrip('.'),
        "valid": False,
        "note": "",
        "detail": ""
    }

    try:
        file_size = os.path.getsize(file_path)
        if file_size == 0:
            result["note"] = "CẢNH BÁO - File rỗng (0 bytes)"
            return result

        with open(file_path, 'rb') as f:
            header = f.read(2048)  # Đọc tối đa 2048 bytes (đủ cho DPX header)

        magic4 = header[:4] if len(header) >= 4 else b''

        # ------------------------------------------------------------------
        # DPX CHECK
        # ------------------------------------------------------------------
        if ext == '.dpx':
            DPX_MAGIC_BE = b'SDPX'  # Big-endian
            DPX_MAGIC_LE = b'XPDS'  # Little-endian

            if magic4 == DPX_MAGIC_BE:
                byte_order = "Big-Endian"
                is_le = False
                result["valid"] = True
            elif magic4 == DPX_MAGIC_LE:
                byte_order = "Little-Endian"
                is_le = True
                result["valid"] = True
            else:
                result["note"] = f"CẢNH BÁO - Không đúng DPX magic (got: {magic4.hex().upper()})"
                return result

            # Parse header fields (nếu đủ dài)
            details = [byte_order]
            if len(header) >= 2048:
                import struct
                endian = '<' if is_le else '>'

                # Offset 0x060 (96): image width (uint32)
                # Offset 0x064 (100): image height (uint32)
                try:
                    width  = struct.unpack_from(f'{endian}I', header, 96)[0]
                    height = struct.unpack_from(f'{endian}I', header, 100)[0]
                    if 0 < width < 100000 and 0 < height < 100000:
                        details.append(f"{width}x{height}px")
                except Exception:
                    pass

                # Offset 0x068 (104): bit depth per channel (uint8) – image element 0
                # Element descriptor at 0x074 (116): uint8
                try:
                    # Image element 0 starts at offset 0x068
                    # Structure per element: 4 bytes data sign, 4 ref low, 4 ref high,
                    #   4 ref low data, 4 ref high data, 1 descriptor, 1 transfer,
                    #   1 colorimetric, 1 bit size ...
                    bit_depth = struct.unpack_from('B', header, 107)[0]  # bit size field
                    if bit_depth in (1, 8, 10, 12, 16, 32, 64):
                        details.append(f"{bit_depth}-bit")
                except Exception:
                    pass

                # Frame rate: offset 0x478 (1144) – ASCII 16 bytes
                try:
                    fps_bytes = header[1144:1160].rstrip(b'\x00').decode('ascii', errors='ignore').strip()
                    if fps_bytes and fps_bytes not in ('', '0', '0.000000'):
                        details.append(f"{fps_bytes} fps")
                except Exception:
                    pass

            result["note"] = "OK - DPX hợp lệ"
            result["detail"] = ", ".join(details)

        # ------------------------------------------------------------------
        # CRI CHECK
        # ------------------------------------------------------------------
        elif ext == '.cri':
            CRI_SIGNATURES = {
                b'CRID': "CRI USM Video",
                b'HCA\x00': "CRI HCA Audio",
                b'@UTF': "CRI UTF Table",
            }

            matched = False
            for sig, label in CRI_SIGNATURES.items():
                if header[:len(sig)] == sig:
                    result["valid"] = True
                    result["note"] = f"OK - {label}"
                    matched = True

                    # Thêm thông tin bổ sung cho USM
                    if sig == b'CRID' and len(header) >= 8:
                        import struct
                        try:
                            chunk_size = struct.unpack_from('>I', header, 4)[0]
                            result["detail"] = f"Chunk size: {chunk_size} bytes"
                        except Exception:
                            pass
                    break

            if not matched:
                # ADX: magic 0x80 0x00 (header[0]=0x80, header[1]=0x00)
                if len(header) >= 2 and header[0] == 0x80 and header[1] == 0x00:
                    result["valid"] = True
                    result["note"] = "OK - CRI ADX Audio"
                    import struct
                    try:
                        copyright_offset = struct.unpack_from('>H', header, 2)[0]
                        result["detail"] = f"Copyright offset: {copyright_offset}"
                    except Exception:
                        pass
                else:
                    result["note"] = f"CẢNH BÁO - Không nhận diện được CRI sub-format (magic: {magic4.hex().upper()})"

        return result

    except PermissionError:
        result["note"] = "CẢNH BÁO - Không có quyền đọc file"
        return result
    except Exception as e:
        result["note"] = f"LỖI - {e}"
        return result


def scan_media_files_logic(root_path, output_file, status_callback, progress_callback,
                           enable_md5=True, enable_ffprobe=True):
    """
    Quét tất cả file media, DPX và CRI.
    enable_md5:     True = tính MD5, False = bỏ qua (nhanh hơn)
    enable_ffprobe: True = lấy thời lượng qua ffprobe, False = bỏ qua (nhanh hơn nhiều)
    Trả về: (message, file_count, folder_summary_dict)
    """

    try:
        # === BƯỚC 1: ĐẾM TỔNG SỐ FILE ===
        status_callback("Đang đếm tổng số file, vui lòng chờ...")
        total_files = 0
        for _, _, files in os.walk(root_path):
            for filename in files:
                if filename.lower().endswith(ALL_EXTENSIONS):
                    total_files += 1

        if total_files == 0:
            return "Hoàn thành! Không tìm thấy file nào.", 0, {}

        # === BƯỚC 2: QUÉT VÀ XỬ LÝ FILE ===
        processed_count = 0
        folder_summary = defaultdict(lambda: defaultdict(int))

        # Batch write: gom dòng rồi flush mỗi 50 file → giảm I/O ghi log
        BATCH_SIZE = 50
        write_buffer = []

        with open(output_file, 'w', encoding='utf-8', buffering=1024*1024) as f:
            if enable_md5:
                f.write("Tên file\tLoại\tDung lượng\tThời lượng\tMD5\tKiểm tra DPX/CRI\tChi tiết\tĐường dẫn\tNgày sửa đổi\tNgày tạo\n")
            else:
                f.write("Tên file\tLoại\tDung lượng\tThời lượng\tKiểm tra DPX/CRI\tChi tiết\tĐường dẫn\tNgày sửa đổi\tNgày tạo\n")

            for root, _, files in os.walk(root_path):
                for filename in files:
                    if not filename.lower().endswith(ALL_EXTENSIONS):
                        continue

                    full_path = os.path.join(root, filename)
                    ext = os.path.splitext(filename)[1].lower()

                    processed_count += 1
                    # Giảm tần suất callback UI: chỉ cập nhật mỗi 5 file
                    # để tránh overhead gọi root.after() liên tục
                    if processed_count % 5 == 0 or processed_count == total_files:
                        status_callback(f"Đang xử lý {processed_count}/{total_files}: {filename}")
                        progress_callback(processed_count, total_files)

                    folder_summary[root][ext] += 1

                    try:
                        # Dùng os.stat() 1 lần thay vì getsize/getmtime/getctime riêng lẻ
                        st = os.stat(full_path)
                        file_size_bytes = st.st_size
                        formatted_size  = format_size(file_size_bytes)
                        mod_time_str    = datetime.fromtimestamp(st.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                        create_time_str = datetime.fromtimestamp(st.st_ctime).strftime('%Y-%m-%d %H:%M:%S')

                        # Tính MD5 (tuỳ chọn) — đọc file 1 lần duy nhất
                        if enable_md5:
                            md5_hash = calculate_md5(full_path)
                        else:
                            md5_hash = None

                        # Xử lý theo loại file
                        if ext in DPX_CRI_EXTENSIONS:
                            check_result    = check_dpx_cri_file(full_path)
                            file_type       = check_result["type"]
                            check_note      = check_result["note"]
                            check_detail    = check_result.get("detail", "")
                            formatted_duration = "N/A"
                        else:
                            file_type    = ext.upper().lstrip('.')
                            check_note   = "-"
                            check_detail = "-"
                            if enable_ffprobe:
                                formatted_duration = format_duration(get_media_duration(full_path))
                            else:
                                formatted_duration = "N/A"

                        if enable_md5:
                            line = (f"{filename}\t{file_type}\t{formatted_size}\t{formatted_duration}\t"
                                    f"{md5_hash}\t{check_note}\t{check_detail}\t"
                                    f"{full_path}\t{mod_time_str}\t{create_time_str}\n")
                        else:
                            line = (f"{filename}\t{file_type}\t{formatted_size}\t{formatted_duration}\t"
                                    f"{check_note}\t{check_detail}\t"
                                    f"{full_path}\t{mod_time_str}\t{create_time_str}\n")

                    except Exception as e:
                        line = f"ERROR__{filename}\tN/A\tN/A\tN/A\tN/A\tLỗi: {e}\t\t{full_path}\tN/A\tN/A\n"

                    write_buffer.append(line)

                    # Flush batch ra đĩa mỗi BATCH_SIZE dòng
                    if len(write_buffer) >= BATCH_SIZE:
                        f.writelines(write_buffer)
                        write_buffer.clear()

            # Flush phần còn lại
            if write_buffer:
                f.writelines(write_buffer)

    except Exception as e:
        return f"Lỗi nghiêm trọng: {e}", 0, {}

    return f"Hoàn thành! Đã quét và ghi lại {processed_count} file.", processed_count, dict(folder_summary)


def get_output_dir():
    """
    Trả về thư mục an toàn để ghi file output, theo từng OS:
    - Windows : thư mục chứa exe / script
    - macOS   : ~/Desktop (thư mục app bị sandbox, không ghi được)
    - Linux   : thư mục chứa script
    """
    if IS_MACOS:
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        if os.path.isdir(desktop):
            return desktop
        return os.path.expanduser("~")  # fallback: home dir
    else:
        # Windows / Linux: ghi cạnh file exe hoặc script
        if getattr(sys, 'frozen', False):
            return os.path.dirname(sys.executable)
        return os.path.dirname(os.path.abspath(__file__)) or os.getcwd()




class MediaScannerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Công cụ quét Media - V2 (DPX/CRI)")
        self.root.geometry("780x560")
        self.root.resizable(True, True)
        self.root.minsize(600, 400)

        style = ttk.Style(self.root)
        # Chọn theme phù hợp từng OS
        if IS_WINDOWS:
            try:
                style.theme_use('vista')
            except Exception:
                style.theme_use('clam')
        elif IS_MACOS:
            try:
                style.theme_use('aqua')
            except Exception:
                style.theme_use('clam')
        else:
            style.theme_use('clam')

        # Notebook (tabs)
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # --- Tab 1: Quét ---
        self.tab_scan = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(self.tab_scan, text="  🔍 Quét File  ")
        self._build_scan_tab()

        # --- Tab 2: Tổng kết ---
        self.tab_summary = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(self.tab_summary, text="  📊 Tổng kết theo Folder  ")
        self._build_summary_tab()

    # =========================================================================
    # BUILD TAB SCAN
    # =========================================================================
    def _build_scan_tab(self):
        frame = self.tab_scan

        ttk.Label(frame, text="Nhập đường dẫn ổ đĩa hoặc thư mục:").grid(
            row=0, column=0, columnspan=2, sticky=tk.W, pady=(0, 4))

        self.path_var = tk.StringVar()
        self.path_entry = ttk.Entry(frame, textvariable=self.path_var, width=70)
        self.path_entry.grid(row=1, column=0, sticky=(tk.W, tk.E), padx=(0, 5))

        self.browse_button = ttk.Button(frame, text="Chọn thư mục...", command=self.browse_folder)
        self.browse_button.grid(row=1, column=1, sticky=tk.E)

        # Loại file cần quét
        ext_frame = ttk.LabelFrame(frame, text="Loại file sẽ được quét", padding="8")
        ext_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=10)

        exts_display = ", ".join([e.upper() for e in MEDIA_EXTENSIONS]) + "   |   DPX, CRI"
        ttk.Label(ext_frame, text=exts_display, foreground="#444", wraplength=650).pack(anchor=tk.W)

        # Tuỳ chọn MD5 + Auto-launch
        opt_frame = ttk.LabelFrame(frame, text="Tuỳ chọn", padding="8")
        opt_frame.grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 6))

        self.md5_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            opt_frame,
            text="Tính MD5 checksum cho mỗi file  (bỏ chọn để quét nhanh hơn với ổ đĩa lớn)",
            variable=self.md5_var
        ).pack(anchor=tk.W)

        self.ffprobe_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            opt_frame,
            text="Lấy thời lượng bằng ffprobe  (bỏ chọn để tăng tốc đáng kể ~200-500ms/file)",
            variable=self.ffprobe_var
        ).pack(anchor=tk.W, pady=(4, 0))

        ttk.Separator(opt_frame, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=6)

        self.launch_excel_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            opt_frame,
            text="Tự động chạy TXT2Excel_media_info_V1.exe sau khi quét xong",
            variable=self.launch_excel_var
        ).pack(anchor=tk.W)

        # Dòng chọn đường dẫn TXT2Excel exe
        exe_row = ttk.Frame(opt_frame)
        exe_row.pack(fill=tk.X, pady=(4, 0))
        ttk.Label(exe_row, text="Đường dẫn EXE:").pack(side=tk.LEFT, padx=(20, 4))
        exe_default = "TXT2Excel_media_info_V1.exe" if IS_WINDOWS else "TXT2Excel_media_info_V1"
        self.excel_exe_var = tk.StringVar(value=exe_default)
        ttk.Entry(exe_row, textvariable=self.excel_exe_var, width=42).pack(side=tk.LEFT)
        ttk.Button(exe_row, text="...", width=3,
                   command=self._browse_excel_exe).pack(side=tk.LEFT, padx=(4, 0))

        self.scan_button = ttk.Button(frame, text="▶  Bắt đầu quét", command=self.start_scan_thread, width=20)
        self.scan_button.grid(row=4, column=0, columnspan=2, pady=8)

        self.progress = ttk.Progressbar(frame, orient=tk.HORIZONTAL, length=100, mode='determinate')
        self.progress.grid(row=5, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 4))

        self.status_var = tk.StringVar(value="Sẵn sàng.")
        status_label = ttk.Label(frame, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W, padding=(4, 2))
        status_label.grid(row=6, column=0, columnspan=2, sticky=(tk.W, tk.E))

        frame.columnconfigure(0, weight=1)

    # =========================================================================
    # BUILD TAB SUMMARY
    # =========================================================================
    def _build_summary_tab(self):
        frame = self.tab_summary

        ttk.Label(frame, text="Tổng kết số lượng file theo từng folder (sau khi quét xong):",
                  font=("Segoe UI", 9, "bold")).pack(anchor=tk.W, pady=(0, 6))

        # Toolbar
        toolbar = ttk.Frame(frame)
        toolbar.pack(fill=tk.X, pady=(0, 6))
        ttk.Button(toolbar, text="🔄 Làm mới", command=self._refresh_summary).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(toolbar, text="💾 Xuất CSV", command=self._export_summary_csv).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(toolbar, text="📊 Xuất Excel", command=self._export_summary_xlsx).pack(side=tk.LEFT)

        self.summary_filter_var = tk.StringVar()
        ttk.Label(toolbar, text="Lọc folder:").pack(side=tk.LEFT, padx=(16, 4))
        filter_entry = ttk.Entry(toolbar, textvariable=self.summary_filter_var, width=30)
        filter_entry.pack(side=tk.LEFT)
        filter_entry.bind("<KeyRelease>", lambda e: self._refresh_summary())

        # Treeview container
        tree_frame = ttk.Frame(frame)
        tree_frame.pack(fill=tk.BOTH, expand=True)

        self._summary_columns = ["Folder", "Tổng", "MP4", "MKV", "AVI", "MOV", "WMV",
                                  "FLV", "WebM", "MPG/MPEG", "MXF", "MP3", "WAV",
                                  "FLAC", "AAC", "OGG", "M4A", "DPX", "CRI"]

        self.summary_tree = ttk.Treeview(
            tree_frame,
            columns=self._summary_columns,
            show="headings",
            selectmode="browse"
        )

        # Column widths
        col_widths = {"Folder": 240, "Tổng": 55}
        for col in self._summary_columns:
            w = col_widths.get(col, 50)
            self.summary_tree.heading(col, text=col,
                                      command=lambda c=col: self._sort_summary(c))
            self.summary_tree.column(col, width=w, minwidth=40, anchor=tk.CENTER if col != "Folder" else tk.W)

        vsb = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.summary_tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=self.summary_tree.xview)
        self.summary_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.summary_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_frame.rowconfigure(0, weight=1)
        tree_frame.columnconfigure(0, weight=1)

        # Footer totals label
        self.summary_total_var = tk.StringVar(value="Chưa có dữ liệu. Hãy thực hiện quét trước.")
        ttk.Label(frame, textvariable=self.summary_total_var, relief=tk.SUNKEN,
                  anchor=tk.W, padding=(4, 2)).pack(fill=tk.X, pady=(6, 0))

        self._folder_summary_data = {}
        self._sort_col = "Folder"
        self._sort_reverse = False

    # =========================================================================
    # SUMMARY LOGIC
    # =========================================================================
    def _populate_summary(self, folder_summary: dict):
        """Nhận dict folder_summary và hiển thị lên Treeview."""
        self._folder_summary_data = folder_summary
        self._refresh_summary()

    def _refresh_summary(self):
        """Vẽ lại bảng tổng kết với filter hiện tại."""
        for row in self.summary_tree.get_children():
            self.summary_tree.delete(row)

        filter_text = self.summary_filter_var.get().lower()
        data = self._folder_summary_data

        ext_map = {
            "MP4": [".mp4"], "MKV": [".mkv"], "AVI": [".avi"], "MOV": [".mov"],
            "WMV": [".wmv"], "FLV": [".flv"], "WebM": [".webm"],
            "MPG/MPEG": [".mpg", ".mpeg"], "MXF": [".mxf"],
            "MP3": [".mp3"], "WAV": [".wav"], "FLAC": [".flac"],
            "AAC": [".aac"], "OGG": [".ogg"], "M4A": [".m4a"],
            "DPX": [".dpx"], "CRI": [".cri"]
        }

        rows = []
        grand_total = 0
        grand_by_col = defaultdict(int)

        for folder_path, ext_counts in data.items():
            if filter_text and filter_text not in folder_path.lower():
                continue

            total = sum(ext_counts.values())
            grand_total += total

            row_vals = [folder_path, total]
            for col in self._summary_columns[2:]:  # skip Folder, Tổng
                cnt = sum(ext_counts.get(e, 0) for e in ext_map.get(col, []))
                row_vals.append(cnt if cnt > 0 else "")
                grand_by_col[col] += cnt

            rows.append(row_vals)

        # Sort
        col_idx = self._summary_columns.index(self._sort_col) if self._sort_col in self._summary_columns else 0
        try:
            rows.sort(key=lambda r: (r[col_idx] if isinstance(r[col_idx], int) else r[col_idx] or ""),
                      reverse=self._sort_reverse)
        except Exception:
            pass

        # Alternating row colors
        for i, row_vals in enumerate(rows):
            tag = "even" if i % 2 == 0 else "odd"
            self.summary_tree.insert("", tk.END, values=row_vals, tags=(tag,))

        self.summary_tree.tag_configure("even", background="#f5f8ff")
        self.summary_tree.tag_configure("odd", background="#ffffff")

        # Update footer
        dpx_total = grand_by_col.get("DPX", 0)
        cri_total = grand_by_col.get("CRI", 0)
        self.summary_total_var.set(
            f"Tổng: {grand_total} file  |  "
            f"Folders: {len(rows)}  |  "
            f"DPX: {dpx_total}  |  CRI: {cri_total}"
        )

        if grand_total == 0 and not self._folder_summary_data:
            self.summary_total_var.set("Chưa có dữ liệu. Hãy thực hiện quét trước.")

    def _sort_summary(self, col):
        if self._sort_col == col:
            self._sort_reverse = not self._sort_reverse
        else:
            self._sort_col = col
            self._sort_reverse = False
        self._refresh_summary()

    def _export_summary_csv(self):
        if not self._folder_summary_data:
            messagebox.showwarning("Chưa có dữ liệu", "Vui lòng thực hiện quét trước khi xuất.")
            return

        save_path = filedialog.asksaveasfilename(
            title="Lưu tổng kết CSV",
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if not save_path:
            return

        try:
            with open(save_path, 'w', encoding='utf-8-sig') as f:
                f.write(",".join(self._summary_columns) + "\n")
                for iid in self.summary_tree.get_children():
                    vals = self.summary_tree.item(iid)["values"]
                    f.write(",".join(str(v) for v in vals) + "\n")
            messagebox.showinfo("Xuất thành công", f"Đã lưu file:\n{save_path}")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể xuất file:\n{e}")

    def _do_export_xlsx(self, save_path, show_dialog=True):
        """Xuất bảng tổng kết ra file Excel (.xlsx) với định dạng màu sắc."""
        # Kiểm tra / cài openpyxl
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            if messagebox.askyesno(
                "Thiếu thư viện",
                "Cần cài openpyxl để xuất Excel.\n\n"
                "Bạn có muốn tự động cài không?\n(Yêu cầu kết nối internet)"
            ):
                import subprocess as _sp
                try:
                    pip_kwargs = {'args': [sys.executable, "-m", "pip", "install", "openpyxl"]}
                    if IS_WINDOWS:
                        pip_kwargs['creationflags'] = _sp.CREATE_NO_WINDOW
                    _sp.check_call(**pip_kwargs)
                    import openpyxl
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    from openpyxl.utils import get_column_letter
                    messagebox.showinfo("Thành công", "Đã cài openpyxl thành công!")
                except Exception as e:
                    messagebox.showerror("Lỗi cài đặt",
                                         f"Không thể cài openpyxl:\n{e}\n\n"
                                         f"Hãy chạy lệnh thủ công:\n  pip install openpyxl")
                    return None
            else:
                return None

        try:
            # ---- Lấy dữ liệu từ Treeview (đã sort/filter) ----
            rows = []
            for iid in self.summary_tree.get_children():
                rows.append(self.summary_tree.item(iid)["values"])

            # ---- Tạo workbook ----
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Tổng kết theo Folder"

            # ---- Màu sắc & style ----
            COLOR_HEADER_BG  = "1F4E79"   # Xanh đậm
            COLOR_HEADER_FG  = "FFFFFF"   # Trắng
            COLOR_TOTAL_BG   = "2E75B6"   # Xanh vừa  (cột Tổng)
            COLOR_MEDIA_BG   = "BDD7EE"   # Xanh nhạt (video/audio)
            COLOR_SPECIAL_BG = "FFE699"   # Vàng nhạt (DPX, CRI)
            COLOR_ROW_EVEN   = "EBF3FB"   # Xanh rất nhạt
            COLOR_ROW_ODD    = "FFFFFF"   # Trắng
            COLOR_FOOTER_BG  = "D6E4F0"   # Xanh footer
            COLOR_WARN_FG    = "C00000"   # Đỏ (giá trị 0)

            def make_fill(hex_color):
                return PatternFill("solid", fgColor=hex_color)

            def make_border():
                thin = Side(style='thin', color="B0C4DE")
                return Border(left=thin, right=thin, top=thin, bottom=thin)

            SPECIAL_COLS = {"DPX", "CRI"}
            TOTAL_COL    = "Tổng"

            # ---- Hàng tiêu đề ----
            ws.append(self._summary_columns)
            header_row = ws[1]
            for i, cell in enumerate(header_row):
                col_name = self._summary_columns[i]
                if col_name in SPECIAL_COLS:
                    cell.fill = make_fill("7030A0")   # Tím cho DPX/CRI
                elif col_name == TOTAL_COL:
                    cell.fill = make_fill(COLOR_TOTAL_BG)
                else:
                    cell.fill = make_fill(COLOR_HEADER_BG)
                cell.font      = Font(bold=True, color=COLOR_HEADER_FG, size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center",
                                           wrap_text=True)
                cell.border    = make_border()
            ws.row_dimensions[1].height = 32

            # ---- Hàng dữ liệu ----
            for r_idx, row_vals in enumerate(rows, start=2):
                ws.append([v for v in row_vals])
                is_even = (r_idx % 2 == 0)
                row_fill = make_fill(COLOR_ROW_EVEN if is_even else COLOR_ROW_ODD)

                for c_idx, cell in enumerate(ws[r_idx], start=1):
                    col_name = self._summary_columns[c_idx - 1]
                    val = row_vals[c_idx - 1]

                    # Nền theo loại cột
                    if col_name in SPECIAL_COLS:
                        cell.fill = make_fill("F3E6FF") if is_even else make_fill("EDD9FF")
                    elif col_name == TOTAL_COL:
                        cell.fill = make_fill("DEEAF1") if is_even else make_fill("C9DEF0")
                    elif col_name == "Folder":
                        cell.fill = row_fill
                    else:
                        cell.fill = row_fill

                    # Căn lề
                    if col_name == "Folder":
                        cell.alignment = Alignment(horizontal="left", vertical="center",
                                                   wrap_text=False)
                    else:
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                    # Màu đỏ nếu giá trị = 0 (ô số, không phải Folder/Tổng)
                    is_numeric_col = col_name not in ("Folder",)
                    if is_numeric_col and str(val) in ("", "0", 0):
                        cell.font = Font(color="BBBBBB", size=10)  # Xám nhạt cho 0/rỗng
                    elif col_name == TOTAL_COL:
                        cell.font = Font(bold=True, size=10)
                    else:
                        cell.font = Font(size=10)

                    cell.border = make_border()

                ws.row_dimensions[r_idx].height = 18

            # ---- Hàng tổng cộng (footer) ----
            footer = []
            for col_name in self._summary_columns:
                if col_name == "Folder":
                    footer.append(f"TỔNG CỘNG  ({len(rows)} folders)")
                else:
                    total = 0
                    for iid in self.summary_tree.get_children():
                        v = self.summary_tree.item(iid)["values"]
                        idx = self._summary_columns.index(col_name)
                        try:
                            total += int(v[idx]) if v[idx] != "" else 0
                        except (ValueError, TypeError):
                            pass
                    footer.append(total if total > 0 else "")
            ws.append(footer)
            footer_row_idx = len(rows) + 2
            for c_idx, cell in enumerate(ws[footer_row_idx], start=1):
                col_name = self._summary_columns[c_idx - 1]
                cell.fill   = make_fill("1F4E79" if col_name in SPECIAL_COLS
                                        else ("2E75B6" if col_name == TOTAL_COL
                                              else COLOR_FOOTER_BG))
                cell.font   = Font(bold=True,
                                   color="FFFFFF" if col_name in (SPECIAL_COLS | {TOTAL_COL})
                                   else "1F3864",
                                   size=10)
                cell.alignment = Alignment(
                    horizontal="left" if col_name == "Folder" else "center",
                    vertical="center"
                )
                cell.border = make_border()
            ws.row_dimensions[footer_row_idx].height = 20

            # ---- Độ rộng cột tự động ----
            col_min_widths = {"Folder": 45, "Tổng": 8}
            for c_idx, col_name in enumerate(self._summary_columns, start=1):
                col_letter = get_column_letter(c_idx)
                min_w = col_min_widths.get(col_name, 7)
                # Đo độ rộng thực tế từ dữ liệu
                max_len = len(col_name)
                for r_idx in range(2, len(rows) + 2):
                    cell_val = str(ws.cell(row=r_idx, column=c_idx).value or "")
                    if len(cell_val) > max_len:
                        max_len = len(cell_val)
                ws.column_dimensions[col_letter].width = max(min_w, min(max_len + 2, 60))

            # ---- Freeze panes (cố định hàng tiêu đề + cột Folder) ----
            ws.freeze_panes = "B2"

            # ---- Auto filter ----
            ws.auto_filter.ref = ws.dimensions

            # ---- Sheet thứ 2: Thống kê tổng hợp ----
            ws2 = wb.create_sheet(title="Thống kê tổng hợp")
            ws2.sheet_view.showGridLines = True

            ext_map = {
                "MP4": [".mp4"], "MKV": [".mkv"], "AVI": [".avi"], "MOV": [".mov"],
                "WMV": [".wmv"], "FLV": [".flv"], "WebM": [".webm"],
                "MPG/MPEG": [".mpg", ".mpeg"], "MXF": [".mxf"],
                "MP3": [".mp3"], "WAV": [".wav"], "FLAC": [".flac"],
                "AAC": [".aac"], "OGG": [".ogg"], "M4A": [".m4a"],
                "DPX": [".dpx"], "CRI": [".cri"]
            }

            ws2.append(["Loại file", "Số lượng", "Tỉ lệ (%)"])
            for cell in ws2[1]:
                cell.fill      = make_fill(COLOR_HEADER_BG)
                cell.font      = Font(bold=True, color="FFFFFF", size=11)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border    = make_border()
            ws2.row_dimensions[1].height = 28

            grand_total_all = 0
            type_counts = {}
            for col_name in self._summary_columns[2:]:
                total = 0
                for folder_counts in self._folder_summary_data.values():
                    for ext in ext_map.get(col_name, []):
                        total += folder_counts.get(ext, 0)
                type_counts[col_name] = total
                grand_total_all += total

            video_types = {"MP4", "MKV", "AVI", "MOV", "WMV", "FLV", "WebM", "MPG/MPEG", "MXF"}
            audio_types = {"MP3", "WAV", "FLAC", "AAC", "OGG", "M4A"}

            for r_idx, (col_name, count) in enumerate(type_counts.items(), start=2):
                pct = f"{count/grand_total_all*100:.1f}%" if grand_total_all > 0 else "0.0%"
                ws2.append([col_name, count, pct])
                is_even = r_idx % 2 == 0
                if col_name in SPECIAL_COLS:
                    bg = "EDD9FF" if is_even else "F3E6FF"
                elif col_name in video_types:
                    bg = "DEEAF1" if is_even else "EBF5FB"
                elif col_name in audio_types:
                    bg = "E2EFDA" if is_even else "F0F7EC"
                else:
                    bg = COLOR_ROW_EVEN if is_even else COLOR_ROW_ODD

                for cell in ws2[r_idx]:
                    cell.fill      = make_fill(bg)
                    cell.alignment = Alignment(horizontal="center" if cell.column > 1 else "left",
                                               vertical="center")
                    cell.font      = Font(size=10,
                                          bold=(col_name in SPECIAL_COLS),
                                          color="7030A0" if col_name in SPECIAL_COLS else "000000")
                    cell.border    = make_border()
                ws2.row_dimensions[r_idx].height = 18

            # Dòng tổng sheet 2
            ws2.append(["TỔNG", grand_total_all, "100%"])
            last_r = len(type_counts) + 2
            for cell in ws2[last_r]:
                cell.fill      = make_fill(COLOR_HEADER_BG)
                cell.font      = Font(bold=True, color="FFFFFF", size=11)
                cell.alignment = Alignment(horizontal="center" if cell.column > 1 else "left",
                                           vertical="center")
                cell.border    = make_border()
            ws2.row_dimensions[last_r].height = 22

            ws2.column_dimensions["A"].width = 16
            ws2.column_dimensions["B"].width = 14
            ws2.column_dimensions["C"].width = 14
            ws2.freeze_panes = "A2"

            # ---- Lưu file ----
            wb.save(save_path)
            if show_dialog:
                messagebox.showinfo(
                    "Xuất Excel thành công",
                    f"Đã lưu file:\n{save_path}\n\n"
                    f"• Sheet 1: Tổng kết theo Folder ({len(rows)} folders)\n"
                    f"• Sheet 2: Thống kê tổng hợp theo loại file"
                )
            return save_path

        except Exception as e:
            messagebox.showerror("Lỗi xuất Excel", f"Không thể xuất file Excel:\n{e}")
            return None

    def _export_summary_xlsx(self):
        """Bấm nút Xuất Excel: hỏi đường dẫn rồi gọi _do_export_xlsx."""
        if not self._folder_summary_data:
            messagebox.showwarning("Chưa có dữ liệu", "Vui lòng thực hiện quét trước khi xuất.")
            return
        save_path = filedialog.asksaveasfilename(
            title="Lưu tổng kết Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if save_path:
            self._do_export_xlsx(save_path, show_dialog=True)

    # =========================================================================
    # SCAN CONTROLS
    # =========================================================================
    def browse_folder(self):
        directory = filedialog.askdirectory(title="Chọn thư mục để quét")
        if directory:
            self.path_var.set(directory)

    def _browse_excel_exe(self):
        """Chọn đường dẫn tới TXT2Excel tool"""
        if IS_WINDOWS:
            filetypes = [("Executable", "*.exe"), ("All files", "*.*")]
        else:
            filetypes = [("All files", "*.*")]
        path = filedialog.askopenfilename(
            title="Chọn TXT2Excel_media_info_V1",
            filetypes=filetypes
        )
        if path:
            self.excel_exe_var.set(path)

    def _launch_txt2excel(self, txt_file_path):
        """
        Gọi TXT2Excel_media_info_V1.exe với đường dẫn file TXT vừa quét.
        Truyền đường dẫn file TXT làm argument dòng lệnh.
        """
        exe_path = self.excel_exe_var.get().strip()

        # Tìm exe: thử đường dẫn tuyệt đối, rồi cùng thư mục script, rồi PATH
        if not os.path.isabs(exe_path):
            # Thử tìm cùng thư mục với script đang chạy
            script_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
            candidate = os.path.join(script_dir, exe_path)
            if os.path.isfile(candidate):
                exe_path = candidate
            elif shutil.which(exe_path):
                exe_path = shutil.which(exe_path)

        if not os.path.isfile(exe_path):
            messagebox.showwarning(
                "Không tìm thấy TXT2Excel",
                f"Không tìm thấy file:\n{exe_path}\n\n"
                f"Vui lòng kiểm tra lại đường dẫn trong mục Tuỳ chọn.\n\n"
                f"File TXT đã được lưu tại:\n{os.path.abspath(txt_file_path)}"
            )
            return

        try:
            abs_txt = os.path.abspath(txt_file_path)
            kwargs = {'args': [exe_path, abs_txt]}
            if IS_WINDOWS:
                kwargs['creationflags'] = subprocess.CREATE_NO_WINDOW
            subprocess.Popen(**kwargs)
            self.update_status(f"Đã khởi chạy TXT2Excel với: {os.path.basename(abs_txt)}")
        except Exception as e:
            messagebox.showerror(
                "Lỗi khởi chạy TXT2Excel",
                f"Không thể chạy:\n{exe_path}\n\nLỗi: {e}"
            )

    def update_status(self, message):
        self.status_var.set(message)

    def update_progress(self, current_value, max_value):
        self.progress['maximum'] = max_value
        self.progress['value'] = current_value

    def scan_completed(self, message, file_count, output_file, folder_summary):
        self.update_status(message)
        self.scan_button.config(state=tk.NORMAL)
        self.browse_button.config(state=tk.NORMAL)
        self.progress['value'] = 0

        # Populate summary tab
        self._populate_summary(folder_summary)

        if file_count > 0:
            # Switch to summary tab to show results
            self.notebook.select(self.tab_summary)
            messagebox.showinfo(
                "Hoàn thành",
                f"{message}\n\nDữ liệu chi tiết đã lưu vào:\n{os.path.abspath(output_file)}\n\n"
                f"Xem tab 'Tổng kết theo Folder' để có cái nhìn tổng quan.\n\n"
                f"⚠ DPX/CRI: kết quả kiểm tra header nằm trong cột 'Kiểm tra DPX/CRI' của file log."
            )

            # Tự động xuất Excel tổng kết (cùng thư mục và tên với file TXT)
            xlsx_path = os.path.splitext(os.path.abspath(output_file))[0] + "_summary.xlsx"
            self.update_status("Đang xuất Excel tổng kết...")
            result = self._do_export_xlsx(xlsx_path, show_dialog=False)
            if result:
                self.update_status(f"Đã xuất Excel: {os.path.basename(xlsx_path)}")
                messagebox.showinfo(
                    "Xuất Excel thành công",
                    f"File tổng kết Excel đã được lưu tự động:\n{xlsx_path}\n\n"
                    f"• Sheet 1: Tổng kết theo Folder\n"
                    f"• Sheet 2: Thống kê tổng hợp theo loại file"
                )

            # Tự động gọi TXT2Excel nếu được chọn
            if self.launch_excel_var.get():
                self._launch_txt2excel(output_file)
        else:
            if "Lỗi nghiêm trọng" in message:
                messagebox.showerror("Lỗi", message)
            else:
                messagebox.showwarning("Hoàn thành",
                                       f"{message}\n\nKhông tìm thấy file nào hoặc đã có lỗi xảy ra.")

    def start_scan_thread(self):
        mapped_drive = self.path_var.get().strip()
        if not os.path.isdir(mapped_drive):
            messagebox.showerror("Lỗi", f"Đường dẫn không hợp lệ:\n'{mapped_drive}'\n\nVui lòng kiểm tra lại.")
            return

        self.scan_button.config(state=tk.DISABLED)
        self.browse_button.config(state=tk.DISABLED)
        self.update_status("Bắt đầu quét, vui lòng chờ...")
        self.progress['value'] = 0

        drive_name = os.path.basename(os.path.normpath(mapped_drive))
        if ":" in drive_name:
            drive_name = drive_name.replace(":", "")
        dt_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename    = f"logfile+{drive_name}_{dt_str}.txt"
        output_filename = os.path.join(get_output_dir(), filename)

        scan_thread = threading.Thread(
            target=self.run_scan_in_background,
            args=(mapped_drive, output_filename, self.md5_var.get(), self.ffprobe_var.get()),
            daemon=True
        )
        scan_thread.start()

    def run_scan_in_background(self, path, output_file, enable_md5, enable_ffprobe):
        progress_callback = lambda current, total: self.root.after(0, self.update_progress, current, total)
        message, file_count, folder_summary = scan_media_files_logic(
            path, output_file, self.update_status, progress_callback,
            enable_md5=enable_md5, enable_ffprobe=enable_ffprobe
        )
        self.root.after(0, self.scan_completed, message, file_count, output_file, folder_summary)


# ==============================================================================
# KIỂM TRA FFPROBE
# ==============================================================================

def check_ffprobe():
    return get_ffprobe_path() is not None


if __name__ == "__main__":
    root = tk.Tk()
    app = MediaScannerApp(root)
    # Nếu ffprobe không tìm thấy (bundled hoặc PATH) → tắt checkbox và cảnh báo
    if not check_ffprobe():
        app.ffprobe_var.set(False)
        messagebox.showwarning(
            "Không tìm thấy FFmpeg",
            "ffprobe không được tìm thấy.\n\n"
            "Option 'Lấy thời lượng bằng ffprobe' đã được tắt tự động.\n"
            "Cột Thời lượng sẽ hiển thị N/A.\n\n"
            "Để bật lại: đặt ffprobe cùng thư mục với chương trình,\n"
            "hoặc cài FFmpeg và thêm vào PATH."
        )
    root.mainloop()
