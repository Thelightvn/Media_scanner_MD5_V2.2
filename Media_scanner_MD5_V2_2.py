import os
import subprocess
import sys
import shutil
import hashlib
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime
from collections import defaultdict


# ==============================================================================
# TỰ ĐỘNG TÌM / CÀI FFPROBE (macOS compatible)
# ==============================================================================

def _find_ffprobe_path():
    """Tìm ffprobe: ưu tiên bundled (PyInstaller), sau đó PATH, cuối cùng imageio-ffmpeg."""
    # 1. Khi đóng gói bằng PyInstaller, ffprobe nằm cạnh executable
    if getattr(sys, 'frozen', False):
        base = sys._MEIPASS if hasattr(sys, '_MEIPASS') else os.path.dirname(sys.executable)
        for name in ('ffprobe', 'ffprobe.exe'):
            candidate = os.path.join(base, name)
            if os.path.isfile(candidate):
                return candidate

    # 2. Thử tìm trong PATH hệ thống
    found = shutil.which("ffprobe")
    if found:
        return found

    # 3. Thử cài imageio-ffmpeg để lấy ffprobe binary
    try:
        import imageio_ffmpeg
        ffmpeg_bin = imageio_ffmpeg.get_ffmpeg_exe()
        ffprobe_bin = ffmpeg_bin.replace("ffmpeg", "ffprobe")
        if os.path.isfile(ffprobe_bin):
            return ffprobe_bin
    except ImportError:
        pass

    # 4. Tự động cài imageio-ffmpeg nếu chưa có
    try:
        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", "imageio-ffmpeg", "--quiet"],
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL
        )
        import imageio_ffmpeg
        ffmpeg_bin = imageio_ffmpeg.get_ffmpeg_exe()
        ffprobe_bin = ffmpeg_bin.replace("ffmpeg", "ffprobe")
        if os.path.isfile(ffprobe_bin):
            return ffprobe_bin
    except Exception:
        pass

    return None  # Không tìm thấy


FFPROBE_PATH = _find_ffprobe_path()


def get_desktop_path():
    """Trả về đường dẫn Desktop của user (macOS / Windows / Linux)."""
    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    if not os.path.isdir(desktop):
        # Fallback về home nếu không có Desktop (headless CI)
        desktop = os.path.expanduser("~")
    return desktop

# ==============================================================================
# PHẦN LOGIC XỬ LÝ
# ==============================================================================

MEDIA_EXTENSIONS = (
    '.mp4', '.mkv', '.avi', '.mov', '.wmv', '.flv', '.webm', '.mpg', '.mpeg', '.mxf',
    '.mp3', '.wav', '.flac', '.aac', '.ogg', '.m4a'
)

DPX_CRI_EXTENSIONS = ('.dpx', '.cri')

ALL_EXTENSIONS = MEDIA_EXTENSIONS + DPX_CRI_EXTENSIONS


def format_size(size_bytes):
    if size_bytes == 0: return "0B"
    size_name = ("B", "KB", "MB", "GB", "TB")
    i = 0
    while size_bytes >= 1024 and i < len(size_name) - 1:
        size_bytes /= 1024.0
        i += 1
    return f"{size_bytes:.2f} {size_name[i]}"


def format_duration(seconds):
    if seconds is None: return "N/A"
    try:
        seconds = float(seconds)
        hours = int(seconds // 3600)
        minutes = int((seconds % 3600) // 60)
        secs = int(seconds % 60)
        return f"{hours:02}:{minutes:02}:{secs:02}"
    except (ValueError, TypeError):
        return "N/A"


def calculate_md5(file_path, chunk_size=8 * 1024 * 1024):
    md5 = hashlib.md5()
    try:
        with open(file_path, 'rb', buffering=0) as f:
            mv = memoryview(bytearray(chunk_size))
            while True:
                n = f.readinto(mv)
                if not n:
                    break
                md5.update(mv[:n])
        return md5.hexdigest().upper()
    except PermissionError:
        return "ERR:NO_PERMISSION"
    except Exception as e:
        return f"ERR:{e}"


def get_media_duration(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    if ext in DPX_CRI_EXTENSIONS:
        return None

    if not FFPROBE_PATH:
        return None

    command = [
        FFPROBE_PATH, "-v", "error", "-show_entries", "format=duration",
        "-of", "default=noprint_wrappers=1:nokey=1", file_path
    ]
    try:
        # Không dùng STARTUPINFO (Windows-only). Ẩn cửa sổ console trên Windows nếu có.
        kwargs = {}
        if sys.platform == "win32":
            si = subprocess.STARTUPINFO()
            si.dwFlags |= subprocess.STARTF_USESHOWWINDOW
            kwargs["startupinfo"] = si

        result = subprocess.run(
            command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True,
            check=False, encoding='utf-8', errors='ignore', **kwargs
        )

        if result.returncode != 0:
            return None
        return result.stdout.strip()

    except FileNotFoundError:
        print(f"CRITICAL ERROR: ffprobe not found at: {FFPROBE_PATH}")
        return None
    except Exception as e:
        print(f"An unexpected error occurred in get_media_duration: {e}")
        return None


def check_dpx_cri_file(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    result = {"type": ext.upper().lstrip('.')}
    return result


def scan_media_files_logic(root_path, output_file, status_callback, progress_callback,
                           enable_md5=True, enable_ffprobe=True):
    try:
        status_callback("Đang đếm tổng số file, vui lòng chờ...")
        total_files = 0
        for _, _, files in os.walk(root_path):
            for filename in files:
                if filename.lower().endswith(ALL_EXTENSIONS):
                    total_files += 1

        if total_files == 0:
            return "Hoàn thành! Không tìm thấy file nào.", 0, {}

        processed_count = 0
        folder_summary = defaultdict(lambda: defaultdict(int))
        BATCH_SIZE = 50
        write_buffer = []

        with open(output_file, 'w', encoding='utf-8', buffering=1024 * 1024) as f:
            f.write("Tên file\tLoại\tDung lượng\tThời lượng\tMD5\tĐường dẫn\n")

            for root, _, files in os.walk(root_path):
                for filename in files:
                    if not filename.lower().endswith(ALL_EXTENSIONS):
                        continue

                    full_path = os.path.join(root, filename)
                    ext = os.path.splitext(filename)[1].lower()

                    processed_count += 1
                    if processed_count % 5 == 0 or processed_count == total_files:
                        status_callback(f"Đang xử lý {processed_count}/{total_files}: {filename}")
                        progress_callback(processed_count, total_files)

                    folder_summary[root][ext] += 1

                    try:
                        st = os.stat(full_path)
                        file_size_bytes = st.st_size
                        formatted_size = format_size(file_size_bytes)

                        md5_hash = calculate_md5(full_path) if enable_md5 else "N/A"

                        if ext in DPX_CRI_EXTENSIONS:
                            file_type = check_dpx_cri_file(full_path)["type"]
                            formatted_duration = "N/A"
                        else:
                            file_type = ext.upper().lstrip('.')
                            if enable_ffprobe:
                                formatted_duration = format_duration(get_media_duration(full_path))
                            else:
                                formatted_duration = "N/A"

                        line = f"{filename}\t{file_type}\t{formatted_size}\t{formatted_duration}\t{md5_hash}\t{full_path}\n"

                    except Exception as e:
                        line = f"ERROR__{filename}\tN/A\tN/A\tN/A\tN/A\t{full_path}\n"

                    write_buffer.append(line)

                    if len(write_buffer) >= BATCH_SIZE:
                        f.writelines(write_buffer)
                        write_buffer.clear()

            if write_buffer:
                f.writelines(write_buffer)

    except Exception as e:
        return f"Lỗi nghiêm trọng: {e}", 0, {}

    return f"Hoàn thành! Đã quét và ghi lại {processed_count} file.", processed_count, dict(folder_summary)


# ==============================================================================
# PHẦN GIAO DIỆN ĐỒ HỌA (GUI)
# ==============================================================================

class MediaScannerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Công cụ quét Media - V2 (DPX/CRI/Auto-Excel)")
        self.root.geometry("780x500")
        self.root.resizable(True, True)
        self.root.minsize(600, 400)

        style = ttk.Style(self.root)
        style.theme_use('vista')

        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        self.tab_scan = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(self.tab_scan, text="  🔍 Quét File  ")
        self._build_scan_tab()

        self.tab_summary = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(self.tab_summary, text="  📊 Tổng kết theo Folder  ")
        self._build_summary_tab()

    def _build_scan_tab(self):
        frame = self.tab_scan

        ttk.Label(frame, text="Nhập đường dẫn ổ đĩa hoặc thư mục:").grid(
            row=0, column=0, columnspan=2, sticky=tk.W, pady=(0, 4))

        self.path_var = tk.StringVar()
        self.path_entry = ttk.Entry(frame, textvariable=self.path_var, width=70)
        self.path_entry.grid(row=1, column=0, sticky=(tk.W, tk.E), padx=(0, 5))

        self.browse_button = ttk.Button(frame, text="Chọn thư mục...", command=self.browse_folder)
        self.browse_button.grid(row=1, column=1, sticky=tk.E)

        ext_frame = ttk.LabelFrame(frame, text="Loại file sẽ được quét", padding="8")
        ext_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=10)

        exts_display = ", ".join([e.upper() for e in MEDIA_EXTENSIONS]) + "   |   DPX, CRI"
        ttk.Label(ext_frame, text=exts_display, foreground="#444", wraplength=650).pack(anchor=tk.W)

        opt_frame = ttk.LabelFrame(frame, text="Tuỳ chọn", padding="8")
        opt_frame.grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 6))

        self.md5_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            opt_frame,
            text="Tính MD5 checksum cho mỗi file  (bỏ chọn để quét nhanh hơn với ổ đĩa lớn)",
            variable=self.md5_var
        ).pack(anchor=tk.W)

        self.ffprobe_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            opt_frame,
            text="Lấy thời lượng bằng ffprobe  (bỏ chọn để tăng tốc đáng kể ~200-500ms/file)",
            variable=self.ffprobe_var
        ).pack(anchor=tk.W, pady=(4, 0))

        self.scan_button = ttk.Button(frame, text="▶  Bắt đầu quét & Xuất Excel", command=self.start_scan_thread,
                                      width=25)
        self.scan_button.grid(row=4, column=0, columnspan=2, pady=12)

        self.progress = ttk.Progressbar(frame, orient=tk.HORIZONTAL, length=100, mode='determinate')
        self.progress.grid(row=5, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 4))

        self.status_var = tk.StringVar(value="Sẵn sàng.")
        status_label = ttk.Label(frame, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W, padding=(4, 2))
        status_label.grid(row=6, column=0, columnspan=2, sticky=(tk.W, tk.E))

        frame.columnconfigure(0, weight=1)

    def _build_summary_tab(self):
        frame = self.tab_summary
        ttk.Label(frame, text="Tổng kết số lượng file theo từng folder (sau khi quét xong):",
                  font=("Segoe UI", 9, "bold")).pack(anchor=tk.W, pady=(0, 6))

        toolbar = ttk.Frame(frame)
        toolbar.pack(fill=tk.X, pady=(0, 6))
        ttk.Button(toolbar, text="🔄 Làm mới", command=self._refresh_summary).pack(side=tk.LEFT, padx=(0, 6))

        self.summary_filter_var = tk.StringVar()
        ttk.Label(toolbar, text="Lọc folder:").pack(side=tk.LEFT, padx=(16, 4))
        filter_entry = ttk.Entry(toolbar, textvariable=self.summary_filter_var, width=30)
        filter_entry.pack(side=tk.LEFT)
        filter_entry.bind("<KeyRelease>", lambda e: self._refresh_summary())

        tree_frame = ttk.Frame(frame)
        tree_frame.pack(fill=tk.BOTH, expand=True)

        self._summary_columns = ["Folder", "Tổng", "MP4", "MKV", "AVI", "MOV", "WMV",
                                 "FLV", "WebM", "MPG/MPEG", "MXF", "MP3", "WAV",
                                 "FLAC", "AAC", "OGG", "M4A", "DPX", "CRI"]

        self.summary_tree = ttk.Treeview(tree_frame, columns=self._summary_columns, show="headings",
                                         selectmode="browse")

        col_widths = {"Folder": 240, "Tổng": 55}
        for col in self._summary_columns:
            w = col_widths.get(col, 50)
            self.summary_tree.heading(col, text=col, command=lambda c=col: self._sort_summary(c))
            self.summary_tree.column(col, width=w, minwidth=40, anchor=tk.CENTER if col != "Folder" else tk.W)

        vsb = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.summary_tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=self.summary_tree.xview)
        self.summary_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.summary_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_frame.rowconfigure(0, weight=1)
        tree_frame.columnconfigure(0, weight=1)

        self.summary_total_var = tk.StringVar(value="Chưa có dữ liệu. Hãy thực hiện quét trước.")
        ttk.Label(frame, textvariable=self.summary_total_var, relief=tk.SUNKEN, anchor=tk.W, padding=(4, 2)).pack(
            fill=tk.X, pady=(6, 0))

        self._folder_summary_data = {}
        self._sort_col = "Folder"
        self._sort_reverse = False

    def _populate_summary(self, folder_summary: dict):
        self._folder_summary_data = folder_summary
        self._refresh_summary()

    def _refresh_summary(self):
        for row in self.summary_tree.get_children():
            self.summary_tree.delete(row)

        filter_text = self.summary_filter_var.get().lower()
        data = self._folder_summary_data

        ext_map = {
            "MP4": [".mp4"], "MKV": [".mkv"], "AVI": [".avi"], "MOV": [".mov"],
            "WMV": [".wmv"], "FLV": [".flv"], "WebM": [".webm"],
            "MPG/MPEG": [".mpg", ".mpeg"], "MXF": [".mxf"],
            "MP3": [".mp3"], "WAV": [".wav"], "FLAC": [".flac"],
            "AAC": [".aac"], "OGG": [".ogg"], "M4A": [".m4a"],
            "DPX": [".dpx"], "CRI": [".cri"]
        }

        rows = []
        grand_total = 0
        grand_by_col = defaultdict(int)

        for folder_path, ext_counts in data.items():
            if filter_text and filter_text not in folder_path.lower():
                continue

            total = sum(ext_counts.values())
            grand_total += total

            row_vals = [folder_path, total]
            for col in self._summary_columns[2:]:
                cnt = sum(ext_counts.get(e, 0) for e in ext_map.get(col, []))
                row_vals.append(cnt if cnt > 0 else "")
                grand_by_col[col] += cnt

            rows.append(row_vals)

        col_idx = self._summary_columns.index(self._sort_col) if self._sort_col in self._summary_columns else 0
        try:
            rows.sort(key=lambda r: (r[col_idx] if isinstance(r[col_idx], int) else r[col_idx] or ""),
                      reverse=self._sort_reverse)
        except Exception:
            pass

        for i, row_vals in enumerate(rows):
            tag = "even" if i % 2 == 0 else "odd"
            self.summary_tree.insert("", tk.END, values=row_vals, tags=(tag,))

        self.summary_tree.tag_configure("even", background="#f5f8ff")
        self.summary_tree.tag_configure("odd", background="#ffffff")

        dpx_total = grand_by_col.get("DPX", 0)
        cri_total = grand_by_col.get("CRI", 0)
        self.summary_total_var.set(
            f"Tổng: {grand_total} file  |  Folders: {len(rows)}  |  DPX: {dpx_total}  |  CRI: {cri_total}")

        if grand_total == 0 and not self._folder_summary_data:
            self.summary_total_var.set("Chưa có dữ liệu. Hãy thực hiện quét trước.")

    def _sort_summary(self, col):
        if self._sort_col == col:
            self._sort_reverse = not self._sort_reverse
        else:
            self._sort_col = col
            self._sort_reverse = False
        self._refresh_summary()

    def _convert_log_to_excel(self, txt_path, folder_summary):
        """Chuyển đổi file log TXT thành file Excel một cách tự động và thêm sheet Tổng hợp"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            try:
                subprocess.check_call(
                    [sys.executable, "-m", "pip", "install", "openpyxl", "--quiet"],
                    stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL
                )
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter
            except Exception as e:
                print(f"Không thể cài đặt openpyxl: {e}")
                return None

        try:
            wb = openpyxl.Workbook()

            # ------------------------------------------------------------------
            # SHEET 1: CHI TIẾT MEDIA
            # ------------------------------------------------------------------
            ws1 = wb.active
            ws1.title = "Chi tiết Media"

            header_fill = PatternFill("solid", fgColor="1F4E79")
            header_font = Font(bold=True, color="FFFFFF")
            thin_border = Border(left=Side(style='thin', color="B0C4DE"),
                                 right=Side(style='thin', color="B0C4DE"),
                                 top=Side(style='thin', color="B0C4DE"),
                                 bottom=Side(style='thin', color="B0C4DE"))

            with open(txt_path, 'r', encoding='utf-8') as f:
                for row_idx, line in enumerate(f, start=1):
                    row_data = line.strip('\n').split('\t')
                    ws1.append(row_data)

                    if row_idx == 1:
                        for cell in ws1[1]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            cell.border = thin_border
                    else:
                        for cell in ws1[row_idx]:
                            cell.border = thin_border
                            if cell.column_letter != 'F':  # Căn giữa trừ cột Đường dẫn
                                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Tự động chỉnh độ rộng cột Sheet 1
            col_widths_s1 = {'A': 30, 'B': 10, 'C': 15, 'D': 15, 'E': 35, 'F': 60}
            for col_letter, width in col_widths_s1.items():
                ws1.column_dimensions[col_letter].width = width

            ws1.freeze_panes = "A2"
            ws1.auto_filter.ref = ws1.dimensions

            # ------------------------------------------------------------------
            # SHEET 2: TỔNG HỢP
            # ------------------------------------------------------------------
            ws2 = wb.create_sheet(title="Tổng hợp")

            summary_columns = ["Folder", "Tổng", "MP4", "MKV", "AVI", "MOV", "WMV",
                               "FLV", "WebM", "MPG/MPEG", "MXF", "MP3", "WAV",
                               "FLAC", "AAC", "OGG", "M4A", "DPX", "CRI"]

            ws2.append(summary_columns)

            # Format header Sheet 2
            for cell in ws2[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            ext_map = {
                "MP4": [".mp4"], "MKV": [".mkv"], "AVI": [".avi"], "MOV": [".mov"],
                "WMV": [".wmv"], "FLV": [".flv"], "WebM": [".webm"],
                "MPG/MPEG": [".mpg", ".mpeg"], "MXF": [".mxf"],
                "MP3": [".mp3"], "WAV": [".wav"], "FLAC": [".flac"],
                "AAC": [".aac"], "OGG": [".ogg"], "M4A": [".m4a"],
                "DPX": [".dpx"], "CRI": [".cri"]
            }

            row_idx = 2
            for folder_path, ext_counts in folder_summary.items():
                total = sum(ext_counts.values())
                row_data = [folder_path, total]

                # Đếm số lượng từng loại theo ext_map
                for col in summary_columns[2:]:
                    cnt = sum(ext_counts.get(e, 0) for e in ext_map.get(col, []))
                    row_data.append(cnt if cnt > 0 else 0)

                ws2.append(row_data)

                # Format data rows
                for c_idx, cell in enumerate(ws2[row_idx], start=1):
                    cell.border = thin_border
                    if c_idx == 1:  # Cột Folder căn trái
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    else:  # Các cột số liệu căn giữa
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        if cell.value == 0:
                            cell.font = Font(color="BBBBBB")  # Làm mờ số 0 cho dễ nhìn
                row_idx += 1

            # Tự động chỉnh độ rộng cột Sheet 2
            ws2.column_dimensions['A'].width = 50
            for i in range(2, len(summary_columns) + 1):
                col_letter = get_column_letter(i)
                ws2.column_dimensions[col_letter].width = 10

            ws2.freeze_panes = "B2"
            ws2.auto_filter.ref = ws2.dimensions

            # ------------------------------------------------------------------
            # LƯU FILE
            # ------------------------------------------------------------------
            xlsx_path = txt_path.replace(".txt", ".xlsx")
            wb.save(xlsx_path)

            # Xóa file txt tạm sau khi xuất excel thành công
            try:
                os.remove(txt_path)
            except:
                pass

            return xlsx_path
        except Exception as e:
            print(f"Lỗi khi lưu file Excel: {e}")
            return None

    def browse_folder(self):
        directory = filedialog.askdirectory(title="Chọn thư mục để quét")
        if directory:
            self.path_var.set(directory)

    def update_status(self, message):
        self.status_var.set(message)

    def update_progress(self, current_value, max_value):
        self.progress['maximum'] = max_value
        self.progress['value'] = current_value

    def scan_completed(self, message, file_count, output_file, folder_summary):
        self.update_status(message)
        self.scan_button.config(state=tk.NORMAL)
        self.browse_button.config(state=tk.NORMAL)
        self.progress['value'] = 0

        self._populate_summary(folder_summary)

        if file_count > 0:
            self.notebook.select(self.tab_summary)

            self.update_status("Đang tự động chuyển đổi dữ liệu sang Excel...")
            excel_path = self._convert_log_to_excel(output_file, folder_summary)

            if excel_path:
                self.update_status(f"Hoàn thành xuất Excel: {os.path.basename(excel_path)}")
                messagebox.showinfo(
                    "Hoàn thành",
                    f"{message}\n\nFile Excel chi tiết đã được tạo tự động tại:\n{os.path.abspath(excel_path)}\n\n"
                    f"File gồm 2 Sheet:\n- Chi tiết Media\n- Tổng hợp số lượng"
                )
            else:
                messagebox.showwarning(
                    "Lỗi xuất Excel",
                    f"Quét thành công nhưng lỗi xuất Excel. Dữ liệu TXT lưu tại:\n{os.path.abspath(output_file)}"
                )
        else:
            if "Lỗi nghiêm trọng" in message:
                messagebox.showerror("Lỗi", message)
            else:
                messagebox.showwarning("Hoàn thành", f"{message}\n\nKhông tìm thấy file nào hoặc đã có lỗi xảy ra.")

    def start_scan_thread(self):
        mapped_drive = self.path_var.get().strip()
        if not os.path.isdir(mapped_drive):
            messagebox.showerror("Lỗi", f"Đường dẫn không hợp lệ:\n'{mapped_drive}'\n\nVui lòng kiểm tra lại.")
            return

        self.scan_button.config(state=tk.DISABLED)
        self.browse_button.config(state=tk.DISABLED)
        self.update_status("Bắt đầu quét, vui lòng chờ...")
        self.progress['value'] = 0

        drive_name = os.path.basename(os.path.normpath(mapped_drive))
        if ":" in drive_name:
            drive_name = drive_name.replace(":", "")
        dt_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"Media_Scan_List_{drive_name}_{dt_str}.txt"

        # Lưu file ra Desktop để tránh lỗi permission
        output_file = os.path.join(get_desktop_path(), output_filename)

        scan_thread = threading.Thread(
            target=self.run_scan_in_background,
            args=(mapped_drive, output_file, self.md5_var.get(), self.ffprobe_var.get()),
            daemon=True
        )
        scan_thread.start()

    def run_scan_in_background(self, path, output_file, enable_md5, enable_ffprobe):
        progress_callback = lambda current, total: self.root.after(0, self.update_progress, current, total)
        message, file_count, folder_summary = scan_media_files_logic(
            path, output_file, self.update_status, progress_callback,
            enable_md5=enable_md5, enable_ffprobe=enable_ffprobe
        )
        self.root.after(0, self.scan_completed, message, file_count, output_file, folder_summary)


def check_ffprobe():
    if FFPROBE_PATH is None:
        ans = messagebox.askyesno(
            "Không tìm thấy FFmpeg",
            "ffprobe không được tìm thấy trong PATH.\n\n"
            "Chương trình vẫn có thể chạy nhưng:\n"
            "  • Cột Thời lượng sẽ hiển thị N/A\n"
            "  • Option 'Lấy thời lượng bằng ffprobe' sẽ bị tắt\n\n"
            "Bạn có muốn tiếp tục không?"
        )
        return ans
    return True


if __name__ == "__main__":
    root = tk.Tk()
    app = MediaScannerApp(root)
    if FFPROBE_PATH is None:
        app.ffprobe_var.set(False)
        messagebox.showwarning(
            "Không tìm thấy FFmpeg",
            "ffprobe không được tìm thấy trong PATH.\n\n"
            "Option 'Lấy thời lượng bằng ffprobe' đã được tắt tự động.\n"
            "Cột Thời lượng sẽ hiển thị N/A.\n\n"
            "Để bật lại: cài FFmpeg và thêm vào PATH, sau đó khởi động lại."
        )
    root.mainloop()